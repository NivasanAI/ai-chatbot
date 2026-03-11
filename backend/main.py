from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
import anthropic
import os
import httpx
from datetime import datetime
from pathlib import Path

app = FastAPI(title="Excel AI Chatbot API with Wikipedia")

# Increase max request size to 20MB for image uploads
from starlette.middleware.base import BaseHTTPMiddleware
class LimitUploadSize(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        request._body = await request.body()
        return await call_next(request)

# CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

import tempfile, pathlib
# On Render the working dir is writable; /tmp is also safe fallback
EXCEL_FILE = str(pathlib.Path(__file__).parent / "chatbot_data.xlsx")
client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY", ""))

# ── Wikipedia helper ───────────────────────────────────────────────────────────

async def search_wikipedia(query: str) -> dict:
    """Search Wikipedia - fetch top 3 results with full content for richer answers."""
    try:
        async with httpx.AsyncClient(timeout=15) as http:
            # Step 1: Search for top 3 matching pages
            search_resp = await http.get(
                "https://en.wikipedia.org/w/api.php",
                params={
                    "action": "query",
                    "list": "search",
                    "srsearch": query,
                    "srlimit": 3,
                    "format": "json",
                },
            )
            search_data = search_resp.json()
            results = search_data.get("query", {}).get("search", [])
            if not results:
                return {"found": False, "title": "", "summary": "", "url": ""}

            # Step 2: Fetch full content for the top result
            top_title = results[0]["title"]
            full_resp = await http.get(
                "https://en.wikipedia.org/w/api.php",
                params={
                    "action": "query",
                    "titles": top_title,
                    "prop": "extracts|info",
                    "exintro": False,        # Get full article not just intro
                    "explaintext": True,     # Plain text no HTML
                    "inprop": "url",
                    "format": "json",
                },
            )
            full_data = full_resp.json()
            pages = full_data.get("query", {}).get("pages", {})
            page = next(iter(pages.values()))
            full_extract = page.get("extract", "")
            page_url = page.get("fullurl", f"https://en.wikipedia.org/wiki/{top_title.replace(' ', '_')}")

            # Trim to ~2000 chars to keep context rich but not too long
            if len(full_extract) > 2000:
                full_extract = full_extract[:2000] + "..."

            # Step 3: Also get summaries of related results (2nd and 3rd)
            related = []
            for r in results[1:3]:
                try:
                    rel_resp = await http.get(
                        f"https://en.wikipedia.org/api/rest_v1/page/summary/{r['title'].replace(' ', '_')}",
                        headers={"Accept": "application/json"},
                    )
                    rel_data = rel_resp.json()
                    rel_extract = rel_data.get("extract", "")
                    if rel_extract:
                        related.append({"title": r["title"], "summary": rel_extract[:400]})
                except Exception:
                    pass

            return {
                "found": True,
                "title": top_title,
                "summary": full_extract,
                "url": page_url,
                "related": related,
            }

    except Exception as e:
        return {"found": False, "title": "", "summary": "", "url": "", "error": str(e)}


# ── Excel helpers ──────────────────────────────────────────────────────────────

def init_excel():
    if Path(EXCEL_FILE).exists():
        # Add WikiSearchLog sheet if missing (for existing files)
        try:
            wb = load_workbook(EXCEL_FILE)
            if "WikiSearchLog" not in wb.sheetnames:
                ws_wiki = wb.create_sheet("WikiSearchLog")
                ws_wiki.append(["session_id", "query", "wikipedia_title", "url", "timestamp"])
                wb.save(EXCEL_FILE)
        except Exception:
            pass
        return

    wb = Workbook()
    ws_chat = wb.active
    ws_chat.title = "ChatHistory"
    ws_chat.append(["id", "session_id", "role", "message", "timestamp"])

    ws_kb = wb.create_sheet("KnowledgeBase")
    ws_kb.append(["topic", "content"])
    ws_kb.append(["company", "We are AcmeCorp, a technology company specialising in AI solutions."])
    ws_kb.append(["hours", "Our support hours are Monday-Friday, 9 AM - 6 PM AEST."])
    ws_kb.append(["contact", "You can reach us at support@acmecorp.com or call 1800-ACME."])

    ws_wiki = wb.create_sheet("WikiSearchLog")
    ws_wiki.append(["session_id", "query", "wikipedia_title", "url", "timestamp"])

    wb.save(EXCEL_FILE)


def read_knowledge_base() -> str:
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name="KnowledgeBase")
        lines = [f"- {row['topic']}: {row['content']}" for _, row in df.iterrows()]
        return "\n".join(lines)
    except Exception:
        return ""


def save_message(session_id: str, role: str, message: str):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["ChatHistory"]
    next_id = ws.max_row
    ws.append([next_id, session_id, role, message, datetime.utcnow().isoformat()])
    wb.save(EXCEL_FILE)


def save_wiki_log(session_id: str, query: str, title: str, url: str):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["WikiSearchLog"]
        ws.append([session_id, query, title, url, datetime.utcnow().isoformat()])
        wb.save(EXCEL_FILE)
    except Exception:
        pass


def read_session_history(session_id: str) -> list[dict]:
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name="ChatHistory")
        session_df = df[df["session_id"] == session_id].sort_values("timestamp")
        return [{"role": row["role"], "content": row["message"]} for _, row in session_df.iterrows()]
    except Exception:
        return []


# ── Models ─────────────────────────────────────────────────────────────────────

class ChatRequest(BaseModel):
    session_id: str
    message: str
    image_base64: str = None
    image_type: str = "image/jpeg"

class ChatResponse(BaseModel):
    reply: str
    session_id: str
    wikipedia_title: str = ""
    wikipedia_url: str = ""

class KnowledgeItem(BaseModel):
    topic: str
    content: str


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.on_event("startup")
async def startup():
    init_excel()


@app.get("/")
def root():
    return {"status": "ok", "message": "Excel AI Chatbot API with Wikipedia is running"}


@app.post("/chat", response_model=ChatResponse)
async def chat(req: ChatRequest):
    if not req.message.strip():
        raise HTTPException(status_code=400, detail="Message cannot be empty")

    save_message(req.session_id, "user", req.message)

    # Fetch Wikipedia data in real-time
    wiki = await search_wikipedia(req.message)

    # Build context
    kb = read_knowledge_base()
    history = read_session_history(req.session_id)

    wiki_context = ""
    if wiki["found"]:
        related_text = ""
        for r in wiki.get("related", []):
            related_text += f"\n\nRelated - {r['title']}:\n{r['summary']}"

        wiki_context = f"""
WIKIPEDIA DATA (real-time):
Main Topic: {wiki['title']}
Source: {wiki['url']}

Full Content:
{wiki['summary']}
{related_text}
"""
        save_wiki_log(req.session_id, req.message, wiki["title"], wiki["url"])

    system_prompt = f"""You are a helpful AI assistant with access to real-time Wikipedia data.

KNOWLEDGE BASE:
{kb if kb else "No custom knowledge base entries."}

{wiki_context}

Instructions:
- Use the Wikipedia data above to give accurate, up-to-date answers
- If Wikipedia data is provided, mention the topic name naturally in your answer
- Always be concise, friendly, and helpful
- If you cite Wikipedia, mention it as your source"""

    # Build message content - support images
    if req.image_base64:
        user_content = [
            {"type": "image", "source": {"type": "base64", "media_type": req.image_type, "data": req.image_base64}},
            {"type": "text", "text": req.message if req.message.strip() else "Analyse this image and estimate the person's age. Describe their approximate age range and key features indicating their age. Note this is an estimate only."}
        ]
        messages = [{"role": "user", "content": user_content}]
        system_prompt = "You are an AI image analysis assistant. When shown a photo of a person, estimate their age range based on visible features like skin, hair, and facial structure. Always clarify this is an AI estimate. Be respectful and professional. If no person is visible, say so politely."
    else:
        messages = history if history else [{"role": "user", "content": req.message}]

    response = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=1000,
        system=system_prompt,
        messages=messages,
    )

    reply = response.content[0].text
    save_message(req.session_id, "assistant", reply)

    return ChatResponse(
        reply=reply,
        session_id=req.session_id,
        wikipedia_title=wiki.get("title", ""),
        wikipedia_url=wiki.get("url", ""),
    )


@app.get("/wikipedia")
async def wikipedia_search(q: str):
    """Directly search Wikipedia - useful for testing."""
    if not q:
        raise HTTPException(status_code=400, detail="Query parameter 'q' is required")
    return await search_wikipedia(q)


@app.get("/wiki-log")
def get_wiki_log():
    """View all Wikipedia searches logged in Excel."""
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name="WikiSearchLog")
        return {"searches": df.to_dict(orient="records")}
    except Exception:
        return {"searches": []}


@app.get("/history/{session_id}")
def get_history(session_id: str):
    return {"session_id": session_id, "messages": read_session_history(session_id)}


@app.get("/knowledge")
def get_knowledge():
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name="KnowledgeBase")
        return {"entries": df.to_dict(orient="records")}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/knowledge")
def add_knowledge(item: KnowledgeItem):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["KnowledgeBase"]
    ws.append([item.topic, item.content])
    wb.save(EXCEL_FILE)
    return {"status": "added", "topic": item.topic}


@app.delete("/knowledge/{topic}")
def delete_knowledge(topic: str):
    df = pd.read_excel(EXCEL_FILE, sheet_name="KnowledgeBase")
    original_len = len(df)
    df = df[df["topic"] != topic]
    if len(df) == original_len:
        raise HTTPException(status_code=404, detail="Topic not found")

    wb = load_workbook(EXCEL_FILE)
    ws = wb["KnowledgeBase"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        ws.cell(row=i, column=1, value=row["topic"])
        ws.cell(row=i, column=2, value=row["content"])
    wb.save(EXCEL_FILE)
    return {"status": "deleted", "topic": topic}


@app.get("/sessions")
def list_sessions():
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name="ChatHistory")
        return {"sessions": df["session_id"].unique().tolist()}
    except Exception:
        return {"sessions": []}
